from bs4 import BeautifulSoup
from urllib.request import urlopen
from pyexcelerate import Workbook
from openpyxl import load_workbook
import time
import sys
import os
from operator import eq
from shutil import copyfile

def Make_Stock_title_xlsx(dir_path, file_name='stock_title'):
    make_xlsx = stock_title()
    make_xlsx.Start_using_this(dir_path=dir_path, file_name=file_name)

def Copy_Stock_title_xlsx(original_path, copy_path, file_name):
    st = stock_title()
    st.dir_path_check(dir_path=copy_path)
    copyfile(original_path + '/' + file_name, copy_path + '/' + file_name)

class stock_title():
    title_list = []

    def Start_using_this(self, dir_path, file_name):            # 최초 실행시 파일을 만들 때 사용합니다.
        self.dir_path_check(dir_path=dir_path)
        self.Get_stock_title()
        self.Check_stock_title_xlsx(dir_path=dir_path, file_name=file_name)

    def Get_html(self, address=None, Error_Count=1):
        try:
            Url = address
            Html = urlopen(Url)
            Source = BeautifulSoup(Html.read(), "html.parser")
            return Source
        except:
            print("Error_Count : " + str(Error_Count) + "\n")
            if Error_Count < 100:
                time.sleep(3)
                return self.Get_html(address, Error_Count + 1)
            else:
                sys.exit(1)

    def Get_stock_title(self):
        # 코스피 부분
        print("Get kospi_title")
        Source1 = self.Get_html("http://finance.naver.com/sise/sise_market_sum.nhn?&page=1")
        time.sleep(3)
        Page_select = Source1.find("table", class_="Nnavi")
        # 1/2/3/4/5/6/7/8/9/10 맨뒤 ... 부분의 코드를 가지게 한다.

        MaxPage_section = Page_select.find("td", class_="pgRR")
        MaxPage_section = MaxPage_section.find("a")
        MaxPage_section = MaxPage_section.get('href')
        MaxPage_section = "http://finance.naver.com" + MaxPage_section
        # 마지막 페이지의 html 주소로 MaxPage_section을 초기화한다.

        Source2 = self.Get_html(MaxPage_section)
        # 1/2/3/4/5 ... 맨뒤 ... 맨뒤로 가기를 선택

        MaxPage_num = Source2.find("td", class_="on")
        MaxPage_num = int(MaxPage_num.find("a").text)
        # 마지막 페이지 번호를 가져온다.

        # 코스피 첫 페이지부터 마지막 페이지까지 기업명과 번호를 가져온다.
        for i in range(1, MaxPage_num + 1, 1):
            Source3 = self.Get_html("http://finance.naver.com/sise/sise_market_sum.nhn?&page=" + str(i))
            # 기업리스트가 있는 페이지를 마지막 페이지까지 차례대로 연다.

            inc_list = Source3.find('table', class_="type_2")
            # 하나의 페이지마다 있는 종목명을 inc_list에 초기화한다.
            inc_list = inc_list.find('tbody')
            inc_list = inc_list.find_all('a', class_="tltle")

            for i in range(0, len(inc_list), 1):
                temp = inc_list[i].text
                self.title_list.append(temp)

        print("Get kosdaq_title")
        ######################################## 코스닥 부분 ###############################################
        Source1 = self.Get_html("http://finance.naver.com/sise/sise_market_sum.nhn?sosok=1")

        Page_select = Source1.find("table", class_="Nnavi")
        # 1/2/3/4/5/6/7/8/9/10 맨뒤 ... 부분의 코드를 가지게 한다.

        MaxPage_section = Page_select.find("td", class_="pgRR")
        MaxPage_section = MaxPage_section.find("a")
        MaxPage_section = MaxPage_section.get('href')
        MaxPage_section = "http://finance.naver.com" + MaxPage_section
        # 마지막 페이지의 html 주소로 MaxPage_section을 초기화한다.

        Source2 = self.Get_html(MaxPage_section)
        # 1/2/3/4/5 ... 맨뒤 ... 맨뒤로 가기를 선택

        MaxPage_num = Source2.find("td", class_="on")
        MaxPage_num = int(MaxPage_num.find("a").text)
        # 마지막 페이지 번호를 가져온다.

        # 코스피 첫 페이지부터 마지막 페이지까지 기업명과 번호를 가져온다.
        for i in range(1, MaxPage_num + 1, 1):
            Source3 = self.Get_html("http://finance.naver.com/sise/sise_market_sum.nhn?sosok=1&page=" + str(i))
            # 기업리스트가 있는 페이지를 마지막 페이지까지 차례대로 연다.

            inc_list = Source3.find('table', class_="type_2")
            # 하나의 페이지마다 있는 종목명을 inc_list에 초기화한다.
            inc_list = inc_list.find('tbody')
            inc_list = inc_list.find_all('a', class_="tltle")

            for i in range(0, len(inc_list), 1):
                temp = inc_list[i].text
                self.title_list.append(temp)

        return self.title_list

    def Check_stock_title_xlsx(self, dir_path, file_name):
        self.dir_path_check(dir_path=dir_path)

        if os.path.exists(dir_path + '/' + file_name):
            print("File already exist")
            self.Update_stock_title_xlsx()
        else:
            print("Make file")
            self.Make_stock_title_xlsx(dir_path, file_name)

    def Make_stock_title_xlsx(self, dir_path, file_name):
        print("Make method")
        wb1 = Workbook()
        ws1 = wb1.new_sheet("sheet1")

        for i in range(0, len(self.title_list), 1):
            ws1.set_cell_value(i + 1, 1, self.title_list[i])

        wb1.save(dir_path + '/' + file_name + '.xlsx')

    def Update_stock_title_xlsx(self, dir_path, file_name):
        comp_list = []

        print("Update method")

        if os.path.isfile(dir_path + '/' + file_name + '.xlsx') == False:
            print("File isn`t exist")
            self.Make_stock_title_xlsx(dir_path, file_name)
            return False

        wb1 = load_workbook(dir_path + '/' + file_name + 'xlsx')
        ws1 = wb1.active

        i = 1

        try:
            while(ws1.cell(row=i, column=1) != '\0'):       #NULL 문자가 아니라면 반복문을 계속 실행
                comp_list.append(ws1.cell(row=i, column=1).value)
                i = i + 1
        except:
            print("Error occured in try-except sentence !")

        for i in range(0, len(self.title_list), 1):
            if eq(comp_list[i], self.title_list[i]) != True :
                print("Update " + dir_path + '/' + file_name + '.xlsx')
                self.Make_stock_title_xlsx(dir_path, file_name)

    def dir_path_check(self, dir_path):
        for i in range(0, len(dir_path) - 1, 1):
            if dir_path[i] == '/' and dir_path[i - 1] != ':':
                dir_name = dir_path[:i]
                if not os.path.isdir(dir_name):
                    os.mkdir(dir_name)

        if not os.path.isdir(dir_path):
            os.mkdir(dir_path)

if __name__ == '__main__':
    test = stock_title()
    test.Start_using_this(dir_path='C:/data/2018_09_29_test', file_name='test1')