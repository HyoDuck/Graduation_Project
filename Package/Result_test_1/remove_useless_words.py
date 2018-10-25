from konlpy.tag import _komoran, Mecab
from operator import eq
from openpyxl import load_workbook
import comp_NaverTrend

def Remove_words(stock_title_dir_path, stock_file_name, keywords, comp_list):
    starter = Word_remover()
    list1, list2 = starter.Remove_words(dir_path=stock_title_dir_path,
                                        file_name=stock_file_name,
                                        keywords=keywords, comp_list=comp_list)

    return list1, list2

class Word_remover:
    keywords_list=[]            # NaverTrend 에서 검색하고 받은 키워드들
    comp_keywords_list = []     # keywords_list와 상반되는 group에서 가져온 단어들
    stock_title_list = []       # 기업명들을 모두 저장한 리스트
    removed_list = []           # 삭제된 키워드들로 초기화할 리스트
    result_list = []            # 최종 결과물

    def Remove_words(self, dir_path, file_name, keywords, comp_list):
        #self.Set_keywords_list(dir_path=keywords_dir_path, array2D_xlsx_path=array2D_xlsx_path, array2D_xlsx_name=array2D_xlsx_name)
        #self.Set_comp_keywords_list(comp_dir_path=comp_dir_path, array2D_xlsx_path=comp_array2D_path, array2D_xlsx_name='comp_' + array2D_xlsx_name)
        self.keywords_list=keywords
        self.comp_keywords_list=comp_list

        print("keywords_list : ")
        print(self.keywords_list)
        print("comp_keywords_list : ")
        print(self.comp_keywords_list)

        self.Set_stock_title_list(dir_path=dir_path, file_name=file_name)

        print("stock_title_list")
        print(self.stock_title_list)

        self.Find_stock_title_words()
        self.Find_normal_words()

        self.list_del_part()

        print("keywords_list : ")
        print(self.keywords_list)
        print("removed_list : ")
        print(self.removed_list)
        print("before return ...")

        return self.keywords_list, self.removed_list


    # keywords_list 를 초기화하는 method입니다.
    def Set_keywords_list(self, dir_path, array2D_xlsx_path, array2D_xlsx_name):
        # comp_NaverTrend.py 의
        # def comp_NaverTrend_xls(dir_path = None, array2D_xlsx_path=None, array2D_xlsx_name='2D_array.xlsx'): 을 call하여 사용합니다.

        self.keywords_list = comp_NaverTrend.comp_NaverTrend_xls(dir_path=dir_path, array2D_xlsx_path=array2D_xlsx_path, array2D_xlsx_name=array2D_xlsx_name)

    # keywords_list 와 상반되는 대조 단어들의 리스트인 comp_keywords_list를 초기화하는 method입니다.
    def Set_comp_keywords_list(self, comp_dir_path, array2D_xlsx_path, array2D_xlsx_name):
        print("Set comp_keywords_list")
        comp_dir_path = comp_dir_path + '/' + 'NaverTrend'

        self.comp_keywords_list = comp_NaverTrend.comp_NaverTrend_xls(dir_path=comp_dir_path, array2D_xlsx_path=array2D_xlsx_path, array2D_xlsx_name=array2D_xlsx_name)

    # stock_title_list의 종목이름들을 초기화할 method 입니다.
    def Set_stock_title_list(self, dir_path, file_name):
        wb1 = load_workbook(dir_path + '/' + file_name + '.xlsx')
        ws1 = wb1.active
        i = 1
        try:
            while(ws1.cell(row=i, column=1).value != None):       #NULL 문자가 아니라면 반복문을 계속 실행
                self.stock_title_list.append(ws1.cell(row=i, column=1).value)
                i = i + 1
        except:
            print("Error occured in try-except sentence !")

    def Find_stock_title_words(self):
        if len(self.keywords_list) == 0 :
            print("plz make keywords_list first")
            return False

        detecter = 0

        for i in range(0, len(self.keywords_list), 1):
            for j in range(0, len(self.stock_title_list), 1):
                if eq(str(self.keywords_list[i]), str(self.stock_title_list[j])):
                    detecter = detecter + 1
                    self.removed_list.append(self.keywords_list[i])

        print(self.removed_list)

                #elif self.keywords_list[i] in self.stock_title_list[j] :
                    #call komoran or Mecab method at this part
    #def Find_including_words(self):
    def Find_normal_words(self):
        for i in range(0, len(self.keywords_list), 1):
            for j in range(0, len(self.comp_keywords_list), 1):
                if eq(self.keywords_list[i], self.comp_keywords_list[j]):
                    self.removed_list.append(self.comp_keywords_list[j])

        print(self.removed_list)

    def list_del_part(self):
        for i in range(0, len(self.removed_list), 1):
            j = 0
            while len(self.keywords_list) != j:
                try:
                    if eq(self.removed_list[i], self.keywords_list[j]):
                        del self.keywords_list[j]
                    else :
                        j = j + 1
                except IndexError as IE :
                    break

if __name__ == '__main__':
    #def Remove_words(stock_title_dir_path, stock_file_name, keywords, comp_list):
    temp_keywords = ['지수', '크리스탈', '아모레퍼시픽', '롯데', '삼성', '삼성전자', 'lg생활건강', '시장', '일본', '중국', '매장', '서울', '크림', '에센스', '마스크', '마스크팩', '팩', '회사', '제품', '대표', '브랜드', '기업', '아모레', 'cj', '토니모리', '신한지주', 'cje&m', '시간', '피부관리']

    #for i in range(0, len(temp_keywords), 1):
    #    temp_keywords[i] = temp_keywords[i].decode('utf-8')

    temp_comps_keywords = ['지수', '크리스탈', '아모레퍼시픽']

    #for i in range(0, len(temp_comps_keywords), 1):
    #    temp_comps_keywords[i] = temp_comps_keywords[i].decode('utf-8')

    Remove_words(stock_title_dir_path='C:/data/Keywords/2018-10-20_01-12-57/Stock_title',
                 stock_file_name='stock_title',
                 keywords=temp_keywords,
                 comp_list=temp_comps_keywords)