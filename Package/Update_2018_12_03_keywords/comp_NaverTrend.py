import os
from openpyxl import load_workbook
import pathlib
from operator import eq
from sklearn.metrics.pairwise import cosine_similarity
import numpy
from pyexcelerate import Workbook

# search_names
# date_group
# num_group     3가지가 각각 검색어에 일치하는 data를 가지고 있음.

def check_date(current_cell_date, past_cell_date) :
    current_year = int(current_cell_date[:4])
    current_month = int(current_cell_date[5:7])
    current_day = int(current_cell_date[8:10])

    past_year = int(past_cell_date[:4])
    past_month = int(past_cell_date[5:7])
    past_day = int(current_cell_date[8:10])

    year = int(current_cell_date[:4]) - int(past_cell_date[:4])
    month = int(current_cell_date[5:7]) - int(past_cell_date[5:7])
    day = int(current_cell_date[8:10]) - int(current_cell_date[8:10])

    #if datetime.date

def get_NaverXls(dir_path, search_name, start_date, end_date):
    real_end_date = end_date
    # real_end_date = end_date[:8]
    # if int(end_date[8:10]) >= 11 :
    #     real_end_date = real_end_date + str(int(end_date[8:10]) - 1)
    # elif int(end_date[8:10]) < 11 :
    #     real_end_date = real_end_date + '0' + str(int(end_date[8:10]) - 1)

    file_path = dir_path + '/' + search_name + '_' + start_date + '_' + end_date + '.xlsx'
    print(file_path)
    wb = load_workbook(file_path)
    ws = wb.active
    date_list = []
    num_list = []

    moving_row = 8
    date_column = 1
    num_column = 2

    while (eq(str(ws.cell(row=moving_row, column=date_column).value), real_end_date) != True) and (ws.cell(row=moving_row, column=date_column).value != None):
        date_list.append(ws.cell(row=moving_row, column=date_column).value)
        num_list.append(ws.cell(row=moving_row, column=num_column).value)
        moving_row = moving_row + 1

    return date_list, num_list

def check_cos_sim(data_num1 = [], data_num2 = []):

    # cosine_similarity에서 사용이 가능하도록 reshape합니다.
    data_num1 = numpy.reshape(data_num1, (1, len(data_num1)))
    data_num2 = numpy.reshape(data_num2, (1, len(data_num2)))

    try:
        result = cosine_similarity(data_num1, data_num2)
    except:
        print("Error occured in cosine_similarity")
        return -1

    return round(float(result), 4)      # 소수점 5자리에서 반올림하고 4자리수까지만 표현하여 return하자

# 2018.09.09 remove_lower_similarity 함수를 만들기로 변경함.
 # def gather_cos_nums(similarity_2D_array, search_names, cos_num_parameter = 0.6):
 #     temp_list = []
 #     result_2D_list = []       # 클러스터링한 결과를 저장할 배열
 # 
 #     for i in range(0, len(search_names), 1):
 #         for j in range(0, len(similarity_2D_array[i]), 1):
 #            if similarity_2D_array[i][j] >= cos_num_parameter:
 #                temp_list.append([])

def calculate_slope(x_peak_pos, y_peak_pos, x_pos, y_pos):
    m = (y_peak_pos - y_pos) / (x_peak_pos - x_pos)

    if m <= 0:
        m = m * (-1)

    return m

def calculate_slope_sum(date_list, num_list):
    sum_of_slopes = 0
    y_pos_peak = 0

    for i in range(0, len(date_list), 1):
        if float(num_list[i]) >= y_pos_peak:
            x_pos_peak = i
            y_pos_peak = 100.0

    for i in range(0, len(date_list), 1):
        if i != x_pos_peak:
            sum_of_slopes = sum_of_slopes + calculate_slope(x_peak_pos=x_pos_peak, y_peak_pos=y_pos_peak, x_pos=i, y_pos=float(num_list[i]))

    return round(sum_of_slopes, 2)

def sorting_slope_sum(slope_sum_list, search_names):
    sorted_sum_list = []

    for i in range(0, len(slope_sum_list), 1):
        sorted_sum_list.append([search_names[i], slope_sum_list[i]])

    for i in range(0, len(slope_sum_list) - 1, 1):
        for j in range(i + 1, len(slope_sum_list), 1):
            if sorted_sum_list[i][1] < sorted_sum_list[j][1]:
                temp = sorted_sum_list[i]
                sorted_sum_list[i] = sorted_sum_list[j]
                sorted_sum_list[j] = temp

    return sorted_sum_list

def make_XlsxFile(similarity_2D_array, search_names, save_dir = 'C:/data/2018_09_09_test/2D_array', save_name = '2D_array.xlsx'):
    wb = Workbook()
    ws = wb.new_sheet('new sheet')

    for i in range(0, len(search_names), 1):
        ws[1][i + 2].value = search_names[i]

    for i in range(0, len(search_names), 1):
        ws[i + 2][1].value = search_names[i]

    for i in range(0, len(search_names), 1):
        for j in range(0, len(search_names), 1):
            ws[i + 2][j + 2].value = similarity_2D_array[i][j]

    wb.save(save_dir + '/' + save_name)

def remove_lower_similarity(average_array, similarity_2D_array, search_names, cos_sim_parameter = 0.70):
    result_list = []
    removing_target_list = []
    removing_num_list = []
    sorted_average_array = []
    average_num = 0

    for i in range(0, len(average_array), 1):
        average_num = average_num + average_array[i]

    average_num = average_num / len(search_names)

    # 평균값과 단어로 배열을 초기화하는 반복문
    for i in range(0, len(search_names), 1):
        sorted_average_array.append([search_names[i], average_array[i]])

    # 평균이 높은순으로 sorting하는 이중반복문
    for i in range(0, len(search_names) - 1, 1):
        for j in range(i + 1, len(search_names), 1):
            if sorted_average_array[i][1] > sorted_average_array[j][1]:
                temp = sorted_average_array[i]
                sorted_average_array[i] = sorted_average_array[j]
                sorted_average_array[j] = temp

    for i in range(0, len(average_array), 1):
        if average_array[i] >= average_num:
            if not search_names[i] in removing_target_list:
                removing_target_list.append(search_names[i])
                removing_num_list.append(i)

    for i in range(0, len(removing_target_list), 1):
        for j in range(0, len(similarity_2D_array), 1):
            if similarity_2D_array[i][j] >= cos_sim_parameter:
                if not search_names[j] in removing_target_list:
                    removing_target_list.append(search_names[j])
                    removing_num_list.append(j)

    # for i in range(0, len(search_names), 1):
    #     if not search_names[i] in removing_target_list:
    #         print(search_names[i])

# 코사인 유사도 분석으로 데이터 리스트의 유사도를 측정하고 수치를 similarity_2D_array 2차원 배열에 append 함.
def make_2D_similarity_array(search_names, num_group, date_group):
    temp_array = []
    similarity_2D_array = []

    for i in range(0, len(search_names), 1):
        for j in range(0, len(search_names), 1):
            temp_array.append(check_cos_sim(num_group[i], num_group[j]))
        similarity_2D_array.append(temp_array)
        temp_array = []

    return similarity_2D_array

def comp_NaverTrend_xls(dir_path = None, array2D_xlsx_path=None, array2D_xlsx_name='2D_array.xlsx', make_2D_array=False):
    filenames = os.listdir(dir_path)
    print(filenames)

    search_names = []

    date_group = []
    num_group = []

    temp_array = []
    similarity_2D_array = []
    average_array = []
    sum_of_slopes_list = []

    # sorted_average_array = []

    # 파일명의 연, 월, 일을 구분하는 부분 시작
    for i in range(0, len(filenames), 1):
        for j in range(0, len(filenames[i]), 1):
            if filenames[i][j] == '_':
                break
        search_names.append(filenames[i][:j])

        if i == 0:
            for k in range(j + 1, len(filenames[i]), 1):
                if filenames[i][k] == '_':
                    break
            start_date = filenames[i][j + 1:k]

            for m in range(k + 1, len(filenames[i]), 1):
                if filenames[i][m] == '.':
                    break
            end_date = filenames[i][k + 1:m]

    print(search_names)
    print(start_date + ' ' + end_date)
    # 파일명의 연, 월, 일을 구분하는 부분 끝.

    # get_NaverXls 함수로 엑셀파일을 열어 데이터를 리스트 형태로 받는다.
    for i in range(0, len(search_names), 1):
        date_list, num_list = get_NaverXls(dir_path=dir_path, search_name=search_names[i], start_date=start_date, end_date=end_date)
        print("successfully loaded")

        date_group.append(date_list)
        num_group.append(num_list)

    # 각각 데이터 리스트의 평균값을 구하는 반복문
    for i in range(0, len(search_names), 1):
        sum_of_nums = 0
        for j in range(0, len(num_group[i]), 1):
            sum_of_nums = sum_of_nums + float(num_group[i][j])

        if sum_of_nums != 0:
            average_array.append(sum_of_nums / len(date_group[i]))
        else:
            average_array.append(0)

    # 데이터 리스트의 최대치를 향한 모든 점의 기울기를 구하고 평균값을 받는다.
    for i in range(0, len(num_group), 1):
        sum_of_slopes_list.append(calculate_slope_sum(date_list=date_group[i], num_list=num_group[i]))

    sorted_slope_sum_list = sorting_slope_sum(sum_of_slopes_list, search_names)
    print(sorted_slope_sum_list)

    # 코사인 유사도 분석으로 데이터 리스트의 유사도를 측정하고 수치를 similarity_2D_array 2차원 배열에 append 함. 2018-10-02 make_2D_array 추가 True 일 때만 생성.
    # 10월 2일 기준으로는 필요없음
    if make_2D_array :
        similarity_2D_array = make_2D_similarity_array(search_names=search_names, num_group=num_group, date_group=date_group)
        make_XlsxFile(similarity_2D_array=similarity_2D_array, search_names=search_names, save_dir=array2D_xlsx_path, save_name=array2D_xlsx_name)

    remove_lower_similarity(average_array=average_array, similarity_2D_array=similarity_2D_array, search_names=search_names)

    return sorted_slope_sum_list

if __name__ == '__main__':
    # get_NaverXls(dir_path, search_name, start_date, end_date)
    print(comp_NaverTrend_xls('C:/data/Keywords_테스트자료2/2018-10-28_23-05-47셀바스헬스케어_포함_2_종목/NaverTrend'))