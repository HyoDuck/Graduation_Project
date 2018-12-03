import openpyxl

def word_list(excel, word, low_num = -0.1, upper_num = 0.1):
    wb = openpyxl.load_workbook(excel)

    ws = wb.active

    word_list = []
    i = 1

    while ws.cell(row=i, column=1).value != word:
        i = i + 1

    print(ws.cell(row=i, column=1).value)

    j = 2
    while ws.cell(row=i, column=j).value != None:
        if ws.cell(row=i, column=j).value >=low_num and ws.cell(row=i, column=j).value <= upper_num:
            word_list.append(ws.cell(row=1, column=j).value)
        j = j + 1

    return word_list

if __name__ == '__main__':
    test = word_list('C:/data/1000.xlsx', '서희건설')
    print(test)
    print(test[:5])