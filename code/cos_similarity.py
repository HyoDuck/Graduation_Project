# Writer : 김효빈
# date : 2018.06.20
# 코사인 유사도를 이용해 특정 종목과 다른 종목들 간 유사도를 측정합니다. 1에 가까울수록 유사도가 높습니다.
# 그래프는 기준이 되는 종목과 이외의 종목 1개씩 일별종가를 이용해 출력합니다.

from sklearn.metrics.pairwise import cosine_similarity
#from sklearn.metrics.pairwise import euclidean_distances
#from sklearn.metrics import jaccard_similarity_score
import numpy
import os
from openpyxl import load_workbook
import pathlib
import matplotlib.pyplot as plt

# 원하는 주식종목의 엑셀파일을 불러오는데 사용하는 함수입니다.
def search(dirname, stock_num):
    filenames = os.listdir(dirname)
    for filename in filenames:
        full_filename = os.path.join(dirname, filename)
        ext = os.path.splitext(full_filename)[0]
        if ext[len(dirname) + 1:len(dirname) + 7] == stock_num:
            print("find : " + ext[len(dirname) + 1:len(dirname) + 7])

            return full_filename

# 원하는 기간동안의 주식종목 종가를 가져옵니다.
def get_price_list(start_date="2018.01.01", end_date="2018.03.01", stock_num="009835"):
    trade_row = 3     # 거래량 탐색 3행부터 시작
    trade_column = 7  # 거래량은 7열에 위치함

    endprice_column = 2 # 종가는 2열에 위치함

    date_row = 3    # 날짜확인 3행부터 시작
    date_column = 1 # 날짜는 1열에 위치함

    start_price = 0;    # start_date에 해당하는 날짜의 가격
    price_dict = {}     # start_date부터 가격을 딕셔너리 형태로 저장함
    price_list = []     # start_date부터 가격을 리스트 형태로 저장함
    avrg_price_list = []

    file_name = search("C:\\data\\StockInfo\\test_list", stock_num)     #디렉토리를 fix시켜놓았기 때문에 이후에 수정이 필요합니다.

    try :
        wb = load_workbook(file_name)
    except :
        return -1

    ws2 = wb.active

    while(1):      # 시작 날짜를 찾음
        try :
            if (ws2.cell(row=date_row + 1, column=date_column).value > start_date) :
                date_row = date_row + 1
                continue
            else :
                break
        except :
            return -1

    start_price = ws2.cell(row=date_row, column=endprice_column).value

    start_price = int(start_price.replace(",", ""))      # 가격에서 , 문자를 제거해 int형으로 변환가능하게 한다.

    while(ws2.cell(row=date_row, column=date_column).value < end_date):    # 끝 날짜까지 반복
        price_dict[ws2.cell(row=date_row, column=date_column).value] = ws2.cell(row=date_row, column=endprice_column).value


        price_list.append(ws2.cell(row=date_row, column=endprice_column).value)
        date_row = date_row - 1

    for i in range(0, len(price_list), 1):
        price_list[i] = int(price_list[i].replace(",", ""))      # 가격 리스트에서 , 문자를 제거해 int형으로 변환 가능하게 합니다.

    # start_date의 가격을 100이라하고 가격변동을 계산하여 리스트로 만듭니다.

    for i in range(0, len(price_list), 1):
        avrg_price_list.append(price_list[i])

    wb.close()

    return avrg_price_list

def check_cos_sim(start_date, end_date, data_num1, data_num2, data1_exist = None):

    # data1_num, data2_num의 일별가격을 가져옵니다.
    # 만약 사용자가 data1의 일별종가 리스트를 data1_exist에 전달인자로 넣어준다면 get_price_list로 또 다시 일별종가를 가져오지 않습니다.
    if data1_exist == None:
        data1 = get_price_list(start_date, end_date, data_num1)
    else:
        data1 = data1_exist
    data2 = get_price_list(start_date, end_date, data_num2)

    if data1 == -1:
        print("data1 doesn`t exist")
        return -1
    elif data2 == -1:
        print("data1 doesn`t exist")
        return -1

    # cosine_similarity에서 사용이 가능하도록 reshape합니다.
    data1 = numpy.reshape(data1, (1, len(data1)))
    data2 = numpy.reshape(data2, (1, len(data2)))

    try:
        result = cosine_similarity(data1, data2)
    except:
        print("Error occured in cosine_similarity")
        return -1

    return result

def kos_cos_sim(start_date, end_date, stock_num ,dir):
    stock_list = get_price_list(start_date, end_date, stock_num)
    temp_list = []
    temp_list.append(stock_num)

    ##### 경로 가져오기
    file_list = list(pathlib.Path(dir).glob('*.xlsx'))  # 경로에서 엑셀파일을 모두 리스트로 초기화

    for i in range(0, len(file_list), 1):  # 절대경로에서 종목번호만 가져오기
        filename_slicing = str(file_list[i])
        file_list[i] = filename_slicing[len(dir) + 1:len(dir) + 7]

    #### 경로 가져오기
    for i in range(0, len(file_list), 1):
        if str(file_list[i]) == stock_num:
            continue

        temp = check_cos_sim(start_date, end_date, stock_num, file_list[i], data1_exist=stock_list)
        if temp == -1:
            continue
        print(temp[0,0])

        if temp[0,0] >= 0.99:       # 정확도를 높이려면 여기를 수정하세요
            temp_list.append(file_list[i])

    return temp_list    # 수정필요

# stock_num_list[0]이 기준으로 사용하였던 주식종목이고 [1] 이후로는 기준과 유사성이 있는 주식종목의 번호들입니다.
# 주식번호를 전달하면 일별종가 그래프로 출력합니다.
def graph_testing(start_date, end_date, stock_num_list):
    print(stock_num_list)
    if len(stock_num_list) < 2:
        print("plz check list")
        return -1

    figure_num = 1

    data1 = get_price_list(start_date, end_date, stock_num_list[0])
    data1_ch = get_price_list(start_date, end_date, stock_num_list[0])

    multinum = data1_ch[0]
    for j in range(0, len(data1), 1):
        data1_ch[j] = data1_ch[j] / multinum

    for i in range(1, len(stock_num_list), 1):
        data2 = get_price_list(start_date, end_date, stock_num_list[i])

        showing(data1, data2, stock_num_list[0], stock_num_list[i], figure_num)
        figure_num = figure_num + 1

        multinum = data2[0]
        for j in range(0, len(data2), 1):
            data2[j] = data2[j] / multinum

        showing(data1_ch, data2, stock_num_list[0], stock_num_list[i], figure_num)
        figure_num = figure_num + 1

def showing(data1, data2, label1, label2, figure_num):
    plt.figure(figure_num)
    plt.plot(data1, label=label1)
    plt.plot(data2, label=label2)

    plt.legend()
    plt.show()

# 테스트
if __name__ == '__main__':
    test_list = kos_cos_sim(start_date="2017.11.01", end_date="2018.06.01", stock_num="014990", dir="C:\\data\\StockInfo\\test_list")
    graph_testing(start_date="2017.11.01", end_date="2018.06.01", stock_num_list=test_list)