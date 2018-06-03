# writer : 김효빈
# last update : 2018.06.03
# 하나의 기준종목과 다른 주식종목들의 유사성을 비교합니다.
# 기준일의 종가를 100이라고 하였을 때 이후의 가격변동을 %로 바꾸어 새로운 배열을 생성하고, 두 개의 배열의 평균이 일정수 이하일 때
# 두 종목은 연관이 있다고 판별합니다.
# comp_kos의 return 값은 (0. 기준종목, 1. 비교종목, 2. 평균치, 3. 기준종목 종가 리스트, 4. 비교종목 종가 리스트) x n 개 입니다.

import math
import pathlib
import glob
from openpyxl import load_workbook
import os

# 클래스 사용하지 않습니다. 2018.06.04 수정
class struct_comp :
    std_stock_num = ''
    comp_stock_num = ''
    comp_list = []
    avrg_result = 0

# 원하는 주식종목의 엑셀파일을 불러오는데 사용하는 함수입니다.
def search(dirname, stock_num):
    filenames = os.listdir(dirname)
    for filename in filenames:
        full_filename = os.path.join(dirname, filename)
        ext = os.path.splitext(full_filename)[0]
        if ext[len(dirname) + 1:len(dirname) + 7] == stock_num:
            print("find : " + ext[len(dirname) + 1:len(dirname) + 7])

            return full_filename

# 원하는 기간동안의 주식종목 종가를 가져옵니다.
def get_price_list(start_date="2018.01.01", end_date="2018.03.01", stock_num="009835"):
    trade_row = 3     # 거래량 탐색 3행부터 시작
    trade_column = 7  # 거래량은 7열에 위치함

    endprice_column = 2 # 종가는 2열에 위치함

    date_row = 3    # 날짜확인 3행부터 시작
    date_column = 1 # 날짜는 1열에 위치함

    start_price = 0;    # start_date에 해당하는 날짜의 가격
    price_dict = {}     # start_date부터 가격을 딕셔너리 형태로 저장함
    price_list = []     # start_date부터 가격을 리스트 형태로 저장함
    avrg_price_list = []

    file_name = search("C:\\data\\StockInfo\\Kospi", stock_num)

    try :
        wb = load_workbook(file_name)
    except :
        return -1
    ws1 = wb.active
    ws2 = wb.active

    while(1):      # 시작 날짜를 찾음
        try :
            if (ws2.cell(row=date_row + 1, column=date_column).value > start_date) :
                date_row = date_row + 1
                continue
            else :
                break
        except :
            return -1

    start_price = ws2.cell(row=date_row, column=endprice_column).value

    start_price = int(start_price.replace(",", ""))      # 가격에서 , 문자를 제거해 int형으로 변환가능하게 한다.

    while(ws2.cell(row=date_row, column=date_column).value < end_date):    # 끝 날짜까지 반복
        price_dict[ws2.cell(row=date_row, column=date_column).value] = ws2.cell(row=date_row, column=endprice_column).value


        price_list.append(ws2.cell(row=date_row, column=endprice_column).value)
        date_row = date_row - 1

    for i in range(0, len(price_list), 1):
        price_list[i] = int(price_list[i].replace(",", ""))      # 가격 리스트에서 , 문자를 제거해 int형으로 변환 가능하게 합니다.

    # start_date의 가격을 100이라하고 가격변동을 계산하여 리스트로 만듭니다.

    for i in range(0, len(price_list), 1):
        avrg_price_list.append((price_list[i] / start_price) * 100)
        avrg_price_list[i] = float("{0:.2f}".format(avrg_price_list[i]))

    wb.close()

    return avrg_price_list

# 100으로 평균화한 2개의 주식종목 리스트를 전달하면 전체 list1 - list2를 return 합니다.
def list_comp(standard_list, compared_list):
    result_list = []

    for i in range(0, len(standard_list), 1):
        result_list.append(standard_list[i] - compared_list[i])
        result_list[i] = float("{0:.2f}".format(result_list[i]))

    return result_list

def get_avrg(extracted_list):
    result = 0

    for i in range(0, len(extracted_list), 1):
        if extracted_list[i] > 0 :
            result = result + extracted_list[i]
        else :
            result = result + (extracted_list[i] * (-1))

    result = result / len(extracted_list)
    result = float("{0:.2f}".format(result))

    return result

def comp_kos(start_date="2018.01.01", end_date="2018.03.01", stock_num="005930", dirname = "C:\\data\\StockInfo\\Kospi", limit = 8):
    std_stock = get_price_list(start_date, end_date, stock_num)

    if std_stock == -1 :
        print("Error occurred in getting std_Abj_close")
        return -1

    temp_struct = struct_comp()
    filename_slicing = ''
    result_list = []
    temp_list = [0, 0, 0, 0, 0]

    # 이차원배열로 초기화해야한다.
    # result_list[n][0]은 기준이 되는 종목번호
    # result_list[n][1]은 비교할 종목번호
    # result_list[n][2]은 차이값 평균치
    # result_list[n][3]은 기준종목의 종가 리스트
    # result_list[n][4]은 비교종목의 종가 리스트

    file_list = list(pathlib.Path("C:\\data\\StockInfo\\Kospi").glob('*.xlsx'))     #경로에서 엑셀파일을 모두 리스트로 초기화

    for i in range(0, len(file_list), 1):           #절대경로에서 종목번호만 가져오기
        filename_slicing = str(file_list[i])
        file_list[i] = filename_slicing[len(dirname) + 1:len(dirname) + 7]

    for i in range(0, len(file_list), 1):
        comp_stock = get_price_list(start_date, end_date, file_list[i])     # 비교할 종목의 일별종가를 리스트로 받아옴.

        if comp_stock == -1 :
            print("Error occurred in getting cmp_Abj_close")
            continue

        #temp_struct.std_stock_num = stock_num
        #temp_struct.comp_stock_num = file_list[i]
        #temp_struct.comp_list = list_comp(std_stock, comp_stock)
        #temp_struct.avrg_result = get_avrg(temp_struct.comp_list)

        temp_list[2] = get_avrg(list_comp(std_stock, comp_stock))       #temp_list[2]를 평균값으로 초기화한다.

        if temp_list[2] > limit :     #평균값이 limit보다 크면 다음 종목번호로 넘어간다.
            continue

        temp_list[0] = stock_num
        temp_list[1] = file_list[i]
        temp_list[3] = std_stock
        temp_list[4] = comp_stock

        print(temp_list)

        result_list.append(temp_list)

    return result_list      # 수정필요

if __name__ == '__main__':
    test_list = comp_kos()